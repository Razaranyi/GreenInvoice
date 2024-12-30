from datetime import datetime
from openpyxl import load_workbook
from logger import Logger

current_date_time = datetime.now().strftime("%Y%m%d_%H%M%S")


class ExcelParser:
    def __init__(self, file_path):
        self.logger = Logger.get_logger(__name__)
        self.data = []
        self.__load_data(file_path)
        self.file_path = file_path

    def __load_data(self, file_path):
        try:
            workbook = load_workbook(filename=file_path)
            sheet = workbook['EFT & Paybox']

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Stop reading when 'Client' is None
                if row[0] is None:
                    break

                row_data = {headers[col_idx]: cell for col_idx, cell in enumerate(row) if col_idx < len(headers)}
                self.data.append(row_data)
        except FileNotFoundError:
            self.logger.error(f"Error: File not found: {file_path}")
            exit(-1)
        except Exception as e:
            self.logger.error(f"An unexpected error occurred: {e}")
            exit(-1)

    def get_row(self, row_index):
        self.logger.debug(f"Getting row {row_index}")
        try:
            if 0 <= row_index < len(self.data):
                return self.data[row_index]
            else:
                raise IndexError(f"Error: Index out of range: {row_index}")
        except IndexError as e:
            self.logger.error(e)
            return None

    def get_cell(self, row_data, column_name):
        try:
            if row_data and column_name in row_data:
                cell_value = row_data[column_name]
                # Check if the cell value is of type float (double) and cast it to int
                if isinstance(cell_value, float):
                    return int(cell_value)
                return cell_value
            else:
                raise KeyError(f"Error: Column not found: {column_name}")
        except KeyError as e:
            self.logger.error(e)
            return None

    def change_invoice_status(self, row_index: int):
        try:
            if 0 <= row_index < len(self.data):
                self.data[row_index]['Invoice'] = True
                self.__save_data()
                self.logger.debug(f"Changed invoice status of row {row_index} to True")
            else:
                raise IndexError(f"Error: Index out of range: {row_index}")
        except IndexError as e:
            self.logger.error(e)

    def __save_data(self):
        try:
            workbook = load_workbook(filename=self.file_path)
            sheet = workbook['EFT & Paybox']

            for idx, row_data in enumerate(self.data):
                for col_idx, (key, value) in enumerate(row_data.items()):
                    sheet.cell(row=idx + 2, column=col_idx + 1, value=value)
            workbook.save(filename=self.file_path)
        except Exception as e:
            self.logger.error(f"An unexpected error occurred while saving: {e}")
